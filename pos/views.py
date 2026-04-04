import json
import csv
import io
import logging
import threading
import re as _re
from decimal import Decimal, InvalidOperation
from datetime import date, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse
from django.db import models, transaction
from django.db.models import Sum, Count, Q, F
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.core.cache import cache
from django.conf import settings

from .decorators import store_required, role_required, superadmin_required
from .utils import store_queryset, get_user_store, log_action
from .models import (
    Store, UserProfile, Product, Category, Supplier, Customer,
    Sale, SaleItem, Return, StoreSettings, TaxRule, Expense,
    ExpenseCategory, SupplierPayment, StockPurchase, StockPurchaseItem,
    Shift, BackupLog, AuditLog,
)

logger = logging.getLogger('pos.security')

# ── Allowed image MIME types ───────────────────────────────────────────────────
_ALLOWED_IMAGE_TYPES = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
_MAX_IMAGE_BYTES      = 3 * 1024 * 1024   # 3 MB

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def get_client_ip(request):
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded:
        return x_forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '127.0.0.1')


def audit(request, action, detail=''):
    """Thin wrapper — delegates to log_action in utils.py."""
    try:
        log_action(request, action, detail)
    except Exception:
        pass


def _safe_redirect(url: str, fallback: str = '/dashboard/') -> str:
    """
    FIX SEC-REDIRECT: Validate 'next' is a relative path.
    Prevents open-redirect attacks by rejecting any URL with a scheme or netloc.
    """
    if not url:
        return fallback
    # Must be a relative path — no scheme, no double-slash host
    if url.startswith('//') or '://' in url or not url.startswith('/'):
        return fallback
    return url


def _validate_image(file_obj) -> str | None:
    """
    FIX SEC-UPLOAD: Validate image MIME type and size.
    Returns an error string on failure, None on success.
    """
    if file_obj.size > _MAX_IMAGE_BYTES:
        return f"Image must be smaller than {_MAX_IMAGE_BYTES // (1024*1024)} MB."
    content_type = file_obj.content_type
    if content_type not in _ALLOWED_IMAGE_TYPES:
        return f"Invalid image type '{content_type}'. Allowed: JPEG, PNG, GIF, WebP."
    return None


def calculate_tax(subtotal, cart_items, store):
    rules = TaxRule.objects.filter(is_active=True)
    if store:
        rules = rules.filter(Q(store=store) | Q(store=None))

    if not rules.exists():
        tax_rate = store.tax_rate if store else Decimal('0')
        return subtotal * (tax_rate / 100), []

    tax_amount = Decimal('0')
    breakdown  = []

    for rule in rules:
        if rule.apply_to == 'all':
            applicable_subtotal = subtotal
        else:
            applicable_subtotal = Decimal('0')
            for item in cart_items:
                try:
                    p = Product.objects.get(id=item['product_id'])
                    if p.category and rule.category and p.category.id == rule.category.id:
                        applicable_subtotal += Decimal(str(item['price'])) * item['quantity']
                except Product.DoesNotExist:
                    pass

        if applicable_subtotal == 0:
            continue

        rate = rule.rate / 100
        if rule.tax_mode == 'exclusive':
            rule_tax = applicable_subtotal * rate
        else:
            rule_tax = applicable_subtotal - (applicable_subtotal / (1 + rate))

        rule_tax = rule_tax.quantize(Decimal('0.01'))
        tax_amount += rule_tax
        breakdown.append({
            'name':   rule.name,
            'type':   rule.tax_type,
            'rate':   float(rule.rate),
            'mode':   rule.tax_mode,
            'amount': float(rule_tax),
        })

    return tax_amount, breakdown


def get_currency(store):
    try:
        s = store.settings
        return s.currency_symbol or 'Rs', s.exchange_rate or Decimal('1')
    except Exception:
        return 'Rs', Decimal('1')


def resolve_store(request):
    """
    Returns a valid store for the request.
    For superusers with no store context, falls back to the first store.
    """
    if request.store:
        return request.store
    if request.user.is_superuser:
        from .models import Store
        return Store.objects.first()
    return None


# ══════════════════════════════════════════════════════════════════════════════
# RATE LIMITING HELPER
# ══════════════════════════════════════════════════════════════════════════════

def _check_login_rate_limit(ip: str, identifier: str, max_attempts: int = 5, window: int = 300) -> bool:
    """
    FIX SEC-RATELIMIT: Simple cache-based rate limiting.
    Returns True if the request is allowed, False if blocked.
    Window is in seconds (default 5 minutes).
    """
    cache_key = f"login_attempts:{identifier}:{ip}"
    attempts  = cache.get(cache_key, 0)
    if attempts >= max_attempts:
        return False
    cache.set(cache_key, attempts + 1, timeout=window)
    return True


def _reset_login_rate_limit(ip: str, identifier: str):
    cache_key = f"login_attempts:{identifier}:{ip}"
    cache.delete(cache_key)


# ══════════════════════════════════════════════════════════════════════════════
# AUTH
# ══════════════════════════════════════════════════════════════════════════════

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        ip       = get_client_ip(request)

        # FIX SEC-RATELIMIT: Block after 5 failed attempts per IP per username
        if not _check_login_rate_limit(ip, username):
            logger.warning(f"Login rate limit hit for username='{username}' ip={ip}")
            messages.error(request, 'Too many login attempts. Please wait 5 minutes.')
            return render(request, 'pos/login.html', {'next': request.GET.get('next', '')})

        user = authenticate(request, username=username, password=password)
        if user:
            _reset_login_rate_limit(ip, username)
            login(request, user)
            audit(request, 'login', f'User "{username}" logged in from {ip}')
            # FIX SEC-REDIRECT: Validate next parameter before redirect
            next_url = _safe_redirect(
                request.POST.get('next') or request.GET.get('next'), '/dashboard/'
            )
            return redirect(next_url)

        audit(request, 'login_failed', f'Failed login for "{username}" from {ip}')
        logger.warning(f"Failed login attempt: username='{username}' ip={ip}")
        messages.error(request, 'Invalid credentials.')

    return render(request, 'pos/login.html', {'next': request.GET.get('next', '')})


@require_POST   # FIX SEC-LOGOUT-CSRF: Logout must be POST to prevent CSRF logout
@login_required
def logout_view(request):
    username = request.user.username
    audit(request, 'logout', f'User "{username}" logged out')
    logout(request)
    return redirect('login')


def pin_login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        pin = request.POST.get('pin', '').strip()
        ip  = get_client_ip(request)

        # FIX SEC-RATELIMIT: PIN brute-force protection (4-digit = 10,000 combinations)
        if not _check_login_rate_limit(ip, 'pin', max_attempts=5, window=300):
            logger.warning(f"PIN login rate limit hit from ip={ip}")
            messages.error(request, 'Too many PIN attempts. Please wait 5 minutes.')
            return render(request, 'pos/pin_login.html', {'next': request.GET.get('next', '')})

        if pin:
            # FIX SEC-PIN: Check hashed PINs, not plaintext
            profile = None
            for p in UserProfile.objects.select_related('user').all():
                if p.check_pin(pin):
                    profile = p
                    break

            if profile:
                _reset_login_rate_limit(ip, 'pin')
                login(request, profile.user,
                      backend='django.contrib.auth.backends.ModelBackend')
                audit(request, 'pin_login',
                      f'PIN login by {profile.user.username} from {ip}')
                next_url = _safe_redirect(
                    request.POST.get('next') or request.GET.get('next'), '/'
                )
                return redirect(next_url)

            audit(request, 'pin_failed', f'Failed PIN login from {ip}')
            logger.warning(f"Failed PIN login from ip={ip}")
            messages.error(request, 'Invalid PIN.')

    return render(request, 'pos/pin_login.html', {'next': request.GET.get('next', '')})


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def dashboard(request):
    store = request.store
    today = timezone.now().date()

    qs = store_queryset(Sale, request).filter(status='completed')

    today_sales   = qs.filter(created_at__date=today)
    today_revenue = today_sales.aggregate(t=Sum('total_amount'))['t'] or 0
    today_count   = today_sales.count()
    week_revenue  = qs.filter(
        created_at__date__gte=today - timedelta(days=7)
    ).aggregate(t=Sum('total_amount'))['t'] or 0

    daily_data = []
    for i in range(6, -1, -1):
        d   = today - timedelta(days=i)
        rev = qs.filter(created_at__date=d).aggregate(t=Sum('total_amount'))['t'] or 0
        daily_data.append({'date': d.strftime('%a'), 'revenue': float(rev)})

    pqs          = store_queryset(Product, request).filter(is_active=True)
    low_stock    = pqs.filter(stock_quantity__lte=F('low_stock_threshold'), stock_quantity__gt=0)
    out_of_stock = pqs.filter(stock_quantity=0)

    sale_ids = qs.values_list('id', flat=True)
    top_products = (
        SaleItem.objects.filter(sale_id__in=sale_ids)
        .values('product_name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:5]
    )
    cat_data = (
        SaleItem.objects.filter(sale_id__in=sale_ids)
        .values('product__category__name')
        .annotate(revenue=Sum('total_price'))
        .order_by('-revenue')[:6]
    )

    return render(request, 'pos/dashboard.html', {
        'today_revenue':      today_revenue,
        'today_count':        today_count,
        'week_revenue':       week_revenue,
        'total_products':     pqs.count(),
        'low_stock_count':    low_stock.count(),
        'out_of_stock_count': out_of_stock.count(),
        'low_stock_items':    low_stock[:5],
        # FIX SEC-XSS: Use json.dumps() output via the |json_script tag in templates,
        # not |safe. We keep json.dumps here but the template must use json_script.
        'daily_data':         json.dumps(daily_data),
        'top_products':       list(top_products),
        'cat_data':           json.dumps([
            {'name': c['product__category__name'] or 'Uncategorized',
             'value': float(c['revenue'])}
            for c in cat_data
        ]),
        'currency_symbol':    get_currency(store)[0],
    })


# ══════════════════════════════════════════════════════════════════════════════
# PRODUCTS
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def product_list(request):
    qs = store_queryset(Product, request).filter(is_active=True).select_related('category', 'supplier')

    search      = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(barcode__icontains=search))
    if category_id:
        qs = qs.filter(category_id=category_id)

    categories = store_queryset(Category, request)
    return render(request, 'pos/products.html', {
        'products':          qs,
        'categories':        categories,
        'search':            search,
        'selected_category': category_id,
    })


@store_required
def product_add(request):
    store = request.store
    if request.method == 'POST':
        try:
            # FIX SEC-INPUT: Validate price/cost fields are valid decimals
            try:
                price      = Decimal(request.POST.get('price', '0'))
                cost_price = Decimal(request.POST.get('cost_price', '0'))
                stock_qty  = int(request.POST.get('stock_quantity', 0))
                threshold  = int(request.POST.get('low_stock_threshold', 10))
            except (InvalidOperation, ValueError):
                messages.error(request, 'Invalid price or quantity values.')
                return redirect('product_add')

            if price < 0 or cost_price < 0 or stock_qty < 0:
                messages.error(request, 'Price, cost and stock cannot be negative.')
                return redirect('product_add')

            name    = request.POST.get('name', '').strip()
            barcode = request.POST.get('barcode', '').strip()
            if not name or not barcode:
                messages.error(request, 'Name and barcode are required.')
                return redirect('product_add')

            p = Product(
                name=name, barcode=barcode,
                price=price, cost_price=cost_price,
                stock_quantity=stock_qty,
                low_stock_threshold=threshold,
                store=store,
            )
            cat_id = request.POST.get('category')
            if cat_id:
                p.category = get_object_or_404(Category, id=cat_id, store=store)
            sup_id = request.POST.get('supplier')
            if sup_id:
                p.supplier = get_object_or_404(Supplier, id=sup_id, store=store)

            if 'image' in request.FILES:
                img_file = request.FILES['image']
                err = _validate_image(img_file)
                if err:
                    messages.error(request, err)
                    return redirect('product_add')
                p.image = img_file

            p.save()
            audit(request, 'product_add',
                  f'Added product "{p.name}" — Barcode: {p.barcode} — Price: {p.price}')
            messages.success(request, 'Product added successfully.')
            return redirect('products')
        except Exception as e:
            logger.error(f"product_add error: {e}")
            messages.error(request, 'An error occurred while adding the product.')

    return render(request, 'pos/product_form.html', {
        'categories': store_queryset(Category, request),
        'suppliers':  store_queryset(Supplier, request),
        'action':     'Add',
    })


@store_required
def product_edit(request, pk):
    store   = request.store
    product = get_object_or_404(Product, pk=pk, store=store)

    if request.method == 'POST':
        try:
            try:
                price      = Decimal(request.POST.get('price', '0'))
                cost_price = Decimal(request.POST.get('cost_price', '0'))
                stock_qty  = int(request.POST.get('stock_quantity', 0))
                threshold  = int(request.POST.get('low_stock_threshold', 10))
            except (InvalidOperation, ValueError):
                messages.error(request, 'Invalid price or quantity values.')
                return redirect('product_edit', pk=pk)

            if price < 0 or cost_price < 0:
                messages.error(request, 'Price and cost cannot be negative.')
                return redirect('product_edit', pk=pk)

            name    = request.POST.get('name', '').strip()
            barcode = request.POST.get('barcode', '').strip()
            if not name or not barcode:
                messages.error(request, 'Name and barcode are required.')
                return redirect('product_edit', pk=pk)

            old_price          = product.price
            product.name       = name
            product.barcode    = barcode
            product.price      = price
            product.cost_price = cost_price
            product.stock_quantity      = stock_qty
            product.low_stock_threshold = threshold

            cat_id = request.POST.get('category')
            product.category = get_object_or_404(Category, id=cat_id, store=store) if cat_id else None

            sup_id = request.POST.get('supplier')
            product.supplier = get_object_or_404(Supplier, id=sup_id, store=store) if sup_id else None

            if 'image' in request.FILES:
                img_file = request.FILES['image']
                err = _validate_image(img_file)
                if err:
                    messages.error(request, err)
                    return redirect('product_edit', pk=pk)
                product.image = img_file

            product.save()

            if str(old_price) != str(product.price):
                audit(request, 'price_change',
                      f'Price of "{product.name}" changed from {old_price} to {product.price}')
            else:
                audit(request, 'product_edit',
                      f'Edited product "{product.name}" — Barcode: {product.barcode}')

            messages.success(request, 'Product updated.')
            return redirect('products')
        except Exception as e:
            logger.error(f"product_edit error: {e}")
            messages.error(request, 'An error occurred while updating the product.')

    return render(request, 'pos/product_form.html', {
        'product':    product,
        'categories': store_queryset(Category, request),
        'suppliers':  store_queryset(Supplier, request),
        'action':     'Edit',
    })


@store_required
@require_POST
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk, store=request.store)
    audit(request, 'product_delete',
          f'Deleted product "{product.name}" — Barcode: {product.barcode}')
    product.is_active = False
    product.save()
    messages.success(request, 'Product removed.')
    return redirect('products')


@store_required
@require_POST
def product_csv_import(request):
    if not request.FILES.get('csv_file'):
        messages.error(request, 'No file provided.')
        return redirect('products')

    store   = request.store
    csv_file = request.FILES['csv_file']

    # FIX SEC-UPLOAD: Validate CSV file type and size
    if csv_file.size > 5 * 1024 * 1024:
        messages.error(request, 'CSV file too large (max 5 MB).')
        return redirect('products')

    try:
        decoded = csv_file.read().decode('utf-8')
    except UnicodeDecodeError:
        messages.error(request, 'File must be UTF-8 encoded.')
        return redirect('products')

    reader = csv.DictReader(io.StringIO(decoded))
    count  = 0
    errors = 0
    for row in reader:
        try:
            barcode = str(row.get('barcode', '')).strip()
            name    = str(row.get('name', '')).strip()
            if not barcode or not name:
                continue
            price = Decimal(str(row.get('price', 0)))
            stock = int(row.get('stock', 0))
            if price < 0 or stock < 0:
                errors += 1
                continue
            Product.objects.update_or_create(
                barcode=barcode, store=store,
                defaults={'name': name, 'price': price, 'stock_quantity': stock, 'store': store}
            )
            count += 1
        except Exception:
            errors += 1

    msg = f'Imported {count} products.'
    if errors:
        msg += f' {errors} rows skipped due to errors.'
    messages.success(request, msg)
    return redirect('products')


# ── Product API endpoints ──────────────────────────────────────────────────────

@login_required   # FIX SEC-AUTH: Was completely unauthenticated
def get_product_by_barcode(request):
    barcode = request.GET.get('barcode', '').strip()
    if not barcode:
        return JsonResponse({'success': False, 'message': 'No barcode provided'})

    qs = Product.objects.filter(barcode=barcode, is_active=True)
    store = get_user_store(request.user)
    if store:
        qs = qs.filter(store=store)
    elif not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'No store assigned'})

    try:
        p = qs.get()
        return JsonResponse({
            'success': True, 'id': p.id, 'name': p.name,
            'price': float(p.price), 'stock': p.stock_quantity,
            'barcode': p.barcode,
            'supplier': p.supplier.name if p.supplier else '',
        })
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Product not found'})


@store_required
def search_products(request):
    q = request.GET.get('q', '').strip()
    if not q or len(q) < 1:
        return JsonResponse({'products': []})

    qs = store_queryset(Product, request).filter(
        is_active=True, stock_quantity__gt=0
    ).filter(
        Q(name__icontains=q) | Q(barcode__icontains=q)
    ).select_related('supplier')

    products = list(qs.values('id', 'name', 'price', 'stock_quantity', 'barcode', 'supplier__name')[:10])
    for p in products:
        p['price']    = float(p['price'])
        p['stock']    = p.pop('stock_quantity')
        p['supplier'] = p.pop('supplier__name') or ''
    return JsonResponse({'products': products})


# ── Categories ─────────────────────────────────────────────────────────────────

@store_required
def category_list(request):
    categories = store_queryset(Category, request)
    return render(request, 'pos/categories.html', {'categories': categories})


@store_required
def category_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Category name is required.')
            return redirect('category_add')
        Category.objects.create(name=name, store=request.store)
        messages.success(request, "Category added successfully")
        return redirect('categories')
    return render(request, 'pos/category_form.html', {'title': 'Add Category'})


@store_required
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk, store=request.store)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Category name is required.')
            return redirect('category_edit', pk=pk)
        category.name = name
        category.save()
        messages.success(request, "Category updated")
        return redirect('categories')
    return render(request, 'pos/category_form.html', {
        'title': 'Edit Category', 'category': category
    })


@store_required
@require_POST   # FIX SEC-CSRF: Category delete must be POST
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk, store=request.store)
    category.delete()
    messages.success(request, "Category deleted")
    return redirect('categories')


# ══════════════════════════════════════════════════════════════════════════════
# CUSTOMER LOYALTY API
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def get_customer_info(request):
    customer_id = request.GET.get('id')
    try:
        c = store_queryset(Customer, request).get(id=customer_id)
        return JsonResponse({
            'success':        True,
            'name':           c.name,
            'loyalty_points': c.loyalty_points,
            'points_value':   float(c.loyalty_points) * 0.5,
        })
    except Customer.DoesNotExist:
        return JsonResponse({'success': False})


# ══════════════════════════════════════════════════════════════════════════════
# POS / CHECKOUT
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def pos_view(request):
    store          = resolve_store(request)
    customers      = store_queryset(Customer, request)
    tax_rules      = store_queryset(TaxRule, request).filter(is_active=True)
    currency_symbol, exchange_rate = get_currency(store)
    return render(request, 'pos/pos.html', {
        'customers':       customers,
        'currency_symbol': currency_symbol,
        'exchange_rate':   exchange_rate,
        'tax_rules':       tax_rules,
    })


@store_required
@require_POST
def checkout(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON'})

    cart               = data.get('cart', [])
    payment_method     = data.get('payment_method', 'cash')
    customer_id        = data.get('customer_id')
    discount_type      = data.get('discount_type', 'none')
    store              = resolve_store(request)

    # FIX SEC-DISCOUNT: Validate discount_type against allowed values
    if discount_type not in ('none', 'percent', 'fixed'):
        return JsonResponse({'success': False, 'message': 'Invalid discount type'})

    if payment_method not in ('cash', 'card', 'online'):
        return JsonResponse({'success': False, 'message': 'Invalid payment method'})

    try:
        amount_received = Decimal(str(data.get('amount_received', 0)))
        discount_value  = Decimal(str(data.get('discount_value', 0)))
        use_loyalty_pts = int(data.get('use_loyalty_points', 0))
    except (InvalidOperation, ValueError):
        return JsonResponse({'success': False, 'message': 'Invalid numeric values'})

    if not cart:
        return JsonResponse({'success': False, 'message': 'Cart is empty'})

    # FIX SEC-PRICE: Validate cart items and fetch prices from the DATABASE
    # Never trust client-submitted prices.
    validated_items = []
    for item in cart:
        try:
            product_id = int(item['product_id'])
            quantity   = int(item['quantity'])
            if quantity <= 0:
                return JsonResponse({'success': False, 'message': 'Quantity must be positive'})
        except (KeyError, ValueError, TypeError):
            return JsonResponse({'success': False, 'message': 'Invalid cart item'})

        try:
            p = store_queryset(Product, request).get(id=product_id, is_active=True)
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Product not found'})

        if p.stock_quantity < quantity:
            return JsonResponse({'success': False,
                                 'message': f'Insufficient stock for {p.name}'})

        # ── Use DB price, not the client-sent price ────────────────────────
        validated_items.append({'product': p, 'quantity': quantity, 'price': p.price})

    subtotal                  = sum(i['price'] * i['quantity'] for i in validated_items)
    tax_amount, tax_breakdown = calculate_tax(
        subtotal,
        [{'product_id': i['product'].id, 'price': float(i['price']), 'quantity': i['quantity']}
         for i in validated_items],
        store
    )

    # Discount
    discount_amount = Decimal('0')
    if discount_type == 'percent' and 0 < discount_value <= 100:
        discount_amount = subtotal * (discount_value / 100)
    elif discount_type == 'fixed' and discount_value > 0:
        discount_amount = min(discount_value, subtotal)

    # Loyalty points
    loyalty_discount   = Decimal('0')
    customer           = None
    final_loyalty_used = 0

    if customer_id:
        try:
            customer       = store_queryset(Customer, request).get(id=customer_id)
            max_points     = min(use_loyalty_pts, customer.loyalty_points)
            loyalty_discount = Decimal(str(max_points)) * Decimal('0.5')
            final_loyalty_used = max_points
        except Customer.DoesNotExist:
            pass

    total  = max(subtotal + tax_amount - discount_amount - loyalty_discount, Decimal('0'))
    change = amount_received - total if payment_method == 'cash' else Decimal('0')
    points_earned = int(total / 100)

    try:
        with transaction.atomic():
            # Re-check stock inside the transaction with select_for_update
            for vi in validated_items:
                p_locked = Product.objects.select_for_update().get(id=vi['product'].id)
                if p_locked.stock_quantity < vi['quantity']:
                    return JsonResponse({'success': False,
                                         'message': f'Stock changed for {p_locked.name}'})

            sale = Sale(
                cashier=request.user,
                store=store,
                subtotal=subtotal,
                tax_amount=tax_amount,
                discount_type=discount_type,
                discount_value=discount_value,
                discount_amount=discount_amount + loyalty_discount,
                loyalty_points_used=final_loyalty_used,
                loyalty_points_earned=points_earned,
                total_amount=total,
                payment_method=payment_method,
                amount_received=amount_received,
                change_amount=change,
            )
            if customer:
                sale.customer = customer
            sale.save()

            if customer:
                customer.loyalty_points = customer.loyalty_points - final_loyalty_used + points_earned
                customer.save()

            for vi in validated_items:
                p = vi['product']
                SaleItem.objects.create(
                    sale=sale, product=p,
                    product_name=p.name, product_barcode=p.barcode,
                    quantity=vi['quantity'],
                    unit_price=vi['price'],
                    total_price=vi['price'] * vi['quantity'],
                )
                Product.objects.filter(id=p.id).update(
                    stock_quantity=F('stock_quantity') - vi['quantity']
                )

    except Exception as e:
        logger.error(f"checkout transaction error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'Checkout error: {e}'})

    audit(request, 'sale_complete',
          f'Sale {sale.sale_number} — {total} — {payment_method} — '
          f'{len(validated_items)} item(s) — Customer: {customer.name if customer else "Walk-in"}')

    _trigger_gdrive_auto_backup(store, f'sale_{sale.sale_number}', 'gdrive_backup_on_sale')

    return JsonResponse({
        'success':         True,
        'sale_id':         sale.id,
        'sale_number':     sale.sale_number,
        'total':           float(total),
        'change':          float(change),
        'points_earned':   points_earned,
        'discount_amount': float(discount_amount + loyalty_discount),
    })


# ══════════════════════════════════════════════════════════════════════════════
# RECEIPTS
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def receipt_view(request, sale_id):
    # Superusers can view any sale; regular users are scoped to their store
    if request.user.is_superuser:
        sale = get_object_or_404(Sale, id=sale_id)
    else:
        sale = get_object_or_404(Sale, id=sale_id, store=request.store)
    store        = sale.store  # always use the sale's own store for settings
    settings_obj, _ = StoreSettings.objects.get_or_create(store=store)

    whatsapp_enabled = (
        settings_obj.whatsapp_enabled
        and bool(settings_obj.whatsapp_token)
        and bool(settings_obj.whatsapp_phone_id)
    )

    currency_symbol, _ = get_currency(store)
    return render(request, 'pos/receipt.html', {
        'sale':             sale,
        'settings':         settings_obj,
        'whatsapp_enabled': whatsapp_enabled,
        'template_name':    settings_obj.whatsapp_template_name or 'receipt',
        'currency_symbol':  currency_symbol,
    })


@store_required
def receipt_pdf(request, sale_id):
    if request.user.is_superuser:
        sale = get_object_or_404(Sale, id=sale_id)
    else:
        sale = get_object_or_404(Sale, id=sale_id, store=request.store)
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT

        buf  = io.BytesIO()
        doc  = SimpleDocTemplate(buf, pagesize=(80*mm, 300*mm),
                                 topMargin=5*mm, bottomMargin=5*mm,
                                 leftMargin=5*mm, rightMargin=5*mm)
        center      = ParagraphStyle('c', alignment=TA_CENTER, fontSize=9)
        right       = ParagraphStyle('r', alignment=TA_RIGHT,  fontSize=9)
        normal      = ParagraphStyle('n', fontSize=8)
        title_style = ParagraphStyle('t', alignment=TA_CENTER, fontSize=13,
                                     fontName='Helvetica-Bold')

        story      = []
        store_name = sale.store.name if sale.store else 'SmartPOS'
        story.append(Paragraph(store_name, title_style))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(f"Receipt #{sale.sale_number}", center))
        story.append(Paragraph(sale.created_at.strftime('%d %b %Y  %H:%M'), center))
        if sale.cashier:
            story.append(Paragraph(
                f"Cashier: {sale.cashier.get_full_name() or sale.cashier.username}", center))
        if sale.customer:
            story.append(Paragraph(f"Customer: {sale.customer.name}", center))
        story.append(HRFlowable(width='100%', thickness=0.5, color=colors.black))
        story.append(Spacer(1, 2*mm))

        for item in sale.items.all():
            story.append(Paragraph(f"<b>{item.product_name}</b>", normal))
            story.append(Paragraph(
                f"  {item.quantity} x Rs {item.unit_price:.2f} = <b>Rs {item.total_price:.2f}</b>",
                normal))

        story.append(Spacer(1, 2*mm))
        story.append(HRFlowable(width='100%', thickness=0.5, color=colors.black))
        story.append(Paragraph(f"Subtotal: Rs {sale.subtotal:.2f}", right))
        if sale.tax_amount > 0:
            story.append(Paragraph(f"Tax: Rs {sale.tax_amount:.2f}", right))
        if sale.discount_amount > 0:
            story.append(Paragraph(f"Discount: - Rs {sale.discount_amount:.2f}", right))
        if sale.loyalty_points_used > 0:
            story.append(Paragraph(f"Points Used: {sale.loyalty_points_used} pts", right))
        story.append(Paragraph(
            f"<b>TOTAL: Rs {sale.total_amount:.2f}</b>",
            ParagraphStyle('bt', alignment=TA_RIGHT, fontSize=11, fontName='Helvetica-Bold')))
        story.append(Paragraph(f"Payment: {sale.get_payment_method_display()}", right))
        if sale.payment_method == 'cash':
            story.append(Paragraph(f"Received: Rs {sale.amount_received:.2f}", right))
            story.append(Paragraph(f"Change: Rs {sale.change_amount:.2f}", right))
        if sale.loyalty_points_earned > 0:
            story.append(Spacer(1, 2*mm))
            story.append(Paragraph(f"Points Earned: +{sale.loyalty_points_earned} pts", center))
            if sale.customer:
                story.append(Paragraph(f"Total Points: {sale.customer.loyalty_points}", center))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph("Thank you for shopping with us!", center))

        doc.build(story)
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="receipt_{sale.sale_number}.pdf"'
        return response

    except ImportError:
        content  = f"RECEIPT #{sale.sale_number}\nDate: {sale.created_at.strftime('%d %b %Y %H:%M')}\n"
        content += "-" * 30 + "\n"
        for item in sale.items.all():
            content += f"{item.product_name} x{item.quantity} = Rs {item.total_price:.2f}\n"
        content += "-" * 30 + "\n"
        if sale.discount_amount > 0:
            content += f"Discount: - Rs {sale.discount_amount:.2f}\n"
        content += f"TOTAL: Rs {sale.total_amount:.2f}\n"
        response = HttpResponse(content, content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="receipt_{sale.sale_number}.txt"'
        return response


@store_required
def thermal_receipt(request, sale_id):
    if request.user.is_superuser:
        sale = get_object_or_404(Sale, id=sale_id)
    else:
        sale = get_object_or_404(Sale, id=sale_id, store=request.store)
    settings_obj, _ = StoreSettings.objects.get_or_create(store=sale.store)
    return render(request, 'pos/thermal_receipt.html', {
        'sale': sale, 'settings': settings_obj,
    })


# ══════════════════════════════════════════════════════════════════════════════
# SALES
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def sales_list(request):
    qs = store_queryset(Sale, request).select_related(
        'cashier', 'customer', 'store'
    ).prefetch_related('items')

    date_from  = request.GET.get('date_from')
    date_to    = request.GET.get('date_to')
    cashier_id = request.GET.get('cashier')
    payment    = request.GET.get('payment')
    search     = request.GET.get('q', '')

    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if cashier_id:
        qs = qs.filter(cashier_id=cashier_id)
    if payment:
        qs = qs.filter(payment_method=payment)
    if search:
        qs = qs.filter(
            Q(sale_number__icontains=search) | Q(customer__name__icontains=search)
        )

    qs = qs.order_by('-created_at')
    cashiers = User.objects.filter(
        profile__store=request.store, sale__isnull=False
    ).distinct()

    return render(request, 'pos/sales.html', {
        'sales':            qs[:100],
        'cashiers':         cashiers,
        'date_from':        date_from,
        'date_to':          date_to,
        'selected_payment': payment,
        'today_str':        date.today().strftime('%Y-%m-%d'),
    })


@store_required
def process_return(request, sale_id):
    if request.user.is_superuser:
        sale = get_object_or_404(Sale, id=sale_id)
    else:
        sale = get_object_or_404(Sale, id=sale_id, store=request.store)
    if request.method == 'POST':
        item_id = request.POST.get('item_id')

        try:
            quantity = int(request.POST.get('quantity', 0))
        except ValueError:
            messages.error(request, 'Invalid quantity.')
            return redirect('process_return', sale_id=sale_id)

        if quantity <= 0:
            messages.error(request, 'Quantity must be positive.')
            return redirect('process_return', sale_id=sale_id)

        reason = request.POST.get('reason', '').strip()
        item   = get_object_or_404(SaleItem, id=item_id, sale=sale)

        if quantity > item.returnable_quantity:
            messages.error(request, 'Return quantity exceeds allowed amount.')
            return redirect('process_return', sale_id=sale_id)

        refund = Decimal(str(item.unit_price)) * quantity
        Return.objects.create(
            sale=sale, sale_item=item,
            quantity_returned=quantity,
            refund_amount=refund,
            reason=reason,
            processed_by=request.user,
        )
        item.returned_quantity += quantity
        item.save()

        if item.product:
            item.product.stock_quantity += quantity
            item.product.save()

        sale.status = (
            'partial_refund'
            if sale.items.filter(returned_quantity__lt=F('quantity')).exists()
            else 'refunded'
        )
        sale.save()

        messages.success(request, f'Return processed. Refund: Rs {refund:.2f}')
        audit(request, 'sale_void',
              f'Return on Sale {sale.sale_number} — Item: {item.product_name} '
              f'x{quantity} — Refund: Rs {refund}')
        return redirect('sales')

    return render(request, 'pos/return_form.html', {'sale': sale})


# ══════════════════════════════════════════════════════════════════════════════
# CUSTOMERS
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def customer_list(request):
    qs     = store_queryset(Customer, request)
    search = request.GET.get('q', '')
    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(phone__icontains=search))
    return render(request, 'pos/customers.html', {'customers': qs, 'search': search})


@store_required
def customer_add(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Customer name is required.')
            return redirect('customer_add')
        Customer.objects.create(
            name=name,
            phone=request.POST.get('phone', ''),
            email=request.POST.get('email', ''),
            address=request.POST.get('address', ''),
            store=request.store,
        )
        messages.success(request, 'Customer added.')
        return redirect('customers')
    return render(request, 'pos/customer_form.html', {'action': 'Add'})


@store_required
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk, store=request.store)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Customer name is required.')
            return redirect('customer_edit', pk=pk)
        customer.name    = name
        customer.phone   = request.POST.get('phone', '')
        customer.email   = request.POST.get('email', '')
        customer.address = request.POST.get('address', '')
        customer.save()
        messages.success(request, 'Customer updated.')
        return redirect('customers')
    return render(request, 'pos/customer_form.html', {'customer': customer, 'action': 'Edit'})


@store_required
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk, store=request.store)
    sales    = store_queryset(Sale, request).filter(customer=customer).order_by('-created_at')
    return render(request, 'pos/customer_detail.html', {'customer': customer, 'sales': sales})


# ══════════════════════════════════════════════════════════════════════════════
# SUPPLIERS
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def supplier_list(request):
    return render(request, 'pos/suppliers.html', {
        'suppliers': store_queryset(Supplier, request)
    })


@store_required
def supplier_add(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Supplier name is required.')
            return redirect('supplier_add')
        Supplier.objects.create(
            name=name,
            phone=request.POST.get('phone', ''),
            email=request.POST.get('email', ''),
            address=request.POST.get('address', ''),
            store=request.store,
        )
        messages.success(request, 'Supplier added.')
        return redirect('suppliers')
    return render(request, 'pos/supplier_form.html', {'action': 'Add'})


@store_required
def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk, store=request.store)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Supplier name is required.')
            return redirect('supplier_edit', pk=pk)
        supplier.name    = name
        supplier.phone   = request.POST.get('phone', '')
        supplier.email   = request.POST.get('email', '')
        supplier.address = request.POST.get('address', '')
        supplier.save()
        messages.success(request, 'Supplier updated.')
        return redirect('suppliers')
    return render(request, 'pos/supplier_form.html', {'supplier': supplier, 'action': 'Edit'})


@store_required
def supplier_payments(request, supplier_id):
    supplier  = get_object_or_404(Supplier, id=supplier_id, store=request.store)
    payments  = SupplierPayment.objects.filter(supplier=supplier).order_by('-created_at')
    purchases = StockPurchase.objects.filter(supplier=supplier).order_by('-created_at')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_payment':
            try:
                amount = Decimal(request.POST['amount'])
                if amount <= 0:
                    raise ValueError
            except (KeyError, InvalidOperation, ValueError):
                messages.error(request, 'Invalid payment amount.')
                return redirect('supplier_payments', supplier_id=supplier_id)

            SupplierPayment.objects.create(
                supplier=supplier, amount=amount,
                payment_method=request.POST.get('payment_method', 'cash'),
                reference=request.POST.get('reference', ''),
                notes=request.POST.get('notes', ''),
                paid_by=request.user,
            )
            supplier.balance_owed = max(Decimal('0'), supplier.balance_owed - amount)
            supplier.save()
            messages.success(request, f'Payment of Rs {amount} recorded.')
            return redirect('supplier_payments', supplier_id=supplier_id)

        elif action == 'add_purchase':
            try:
                paid = Decimal(request.POST.get('amount_paid', 0))
            except InvalidOperation:
                paid = Decimal('0')

            product_ids   = request.POST.getlist('product_id[]')
            quantities    = request.POST.getlist('quantity[]')
            unit_costs    = request.POST.getlist('unit_cost[]')
            product_names = request.POST.getlist('product_name[]')

            total      = Decimal('0')
            line_items = []
            for pid, qty, cost, pname in zip(product_ids, quantities, unit_costs, product_names):
                try:
                    qty_int  = int(qty)
                    cost_dec = Decimal(str(cost))
                    if qty_int <= 0 or cost_dec < 0:
                        continue
                    line_total = qty_int * cost_dec
                    total     += line_total
                    line_items.append({
                        'product_id':   int(pid) if pid else None,
                        'product_name': pname,
                        'quantity':     qty_int,
                        'unit_cost':    cost_dec,
                        'total_cost':   line_total,
                    })
                except Exception:
                    continue

            if not line_items:
                total = Decimal(str(request.POST.get('total_amount', 0)))

            purchase = StockPurchase.objects.create(
                supplier=supplier, store=request.store,
                total_amount=total, amount_paid=paid,
                notes=request.POST.get('notes', ''),
                added_by=request.user,
            )

            for item in line_items:
                product = None
                if item['product_id']:
                    try:
                        product = store_queryset(Product, request).get(id=item['product_id'])
                        product.stock_quantity += item['quantity']
                        product.cost_price      = item['unit_cost']
                        product.save()
                    except Product.DoesNotExist:
                        pass

                StockPurchaseItem.objects.create(
                    purchase=purchase, product=product,
                    product_name=item['product_name'],
                    quantity=item['quantity'],
                    unit_cost=item['unit_cost'],
                    total_cost=item['total_cost'],
                )

            supplier.balance_owed += (total - paid)
            supplier.save()
            audit(request, 'stock_purchase',
                  f'Stock purchase from {supplier.name} — Total: Rs {total} — '
                  f'Paid: Rs {paid} — {len(line_items)} item(s)')
            messages.success(request, f'Stock purchase recorded.')
            return redirect('supplier_payments', supplier_id=supplier_id)

    return render(request, 'pos/supplier_payments.html', {
        'supplier':  supplier,
        'payments':  payments,
        'purchases': purchases,
        'products':  store_queryset(Product, request).filter(is_active=True),
    })


# ══════════════════════════════════════════════════════════════════════════════
# ANALYTICS & REPORTS
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def analytics(request):
    today  = date.today()
    # FIX SEC-INPUT: Clamp period to safe range
    try:
        days = max(1, min(int(request.GET.get('period', '30')), 365))
    except ValueError:
        days = 30
    start_date = today - timedelta(days=days)

    qs = store_queryset(Sale, request).filter(
        status='completed', created_at__date__gte=start_date
    )
    sale_ids = qs.values_list('id', flat=True)

    daily = []
    for i in range(days - 1, -1, -1):
        d   = today - timedelta(days=i)
        rev = qs.filter(created_at__date=d).aggregate(t=Sum('total_amount'))['t'] or 0
        daily.append({'date': d.strftime('%d %b'), 'revenue': float(rev)})

    top_products = (
        SaleItem.objects.filter(sale_id__in=sale_ids)
        .values('product_name')
        .annotate(units=Sum('quantity'), revenue=Sum('total_price'))
        .order_by('-units')[:10]
    )
    cat_sales = (
        SaleItem.objects.filter(sale_id__in=sale_ids)
        .values('product__category__name')
        .annotate(revenue=Sum('total_price'))
        .order_by('-revenue')
    )

    total_revenue      = qs.aggregate(t=Sum('total_amount'))['t'] or 0
    total_transactions = qs.count()

    return render(request, 'pos/analytics.html', {
        'daily_data':         json.dumps(daily),
        'top_products':       list(top_products),
        'cat_sales':          json.dumps([
            {'name': c['product__category__name'] or 'Uncategorized',
             'value': float(c['revenue'])}
            for c in cat_sales
        ]),
        'total_revenue':      total_revenue,
        'total_transactions': total_transactions,
        'avg_transaction':    total_revenue / total_transactions if total_transactions else 0,
        'period':             str(days),
    })


@store_required
def profit_report(request):
    today  = date.today()
    try:
        days = max(1, min(int(request.GET.get('period', '30')), 365))
    except ValueError:
        days = 30
    start_date = today - timedelta(days=days)

    sale_qs = store_queryset(Sale, request).filter(
        status='completed', created_at__date__gte=start_date
    )
    items         = SaleItem.objects.filter(sale__in=sale_qs).select_related('product')
    total_revenue = sale_qs.aggregate(t=Sum('total_amount'))['t'] or 0
    total_cost    = sum(
        float(i.product.cost_price if i.product else 0) * i.quantity for i in items
    )
    total_profit  = float(total_revenue) - total_cost
    profit_margin = (total_profit / float(total_revenue) * 100) if total_revenue else 0

    daily_data = []
    for i in range(days - 1, -1, -1):
        d          = today - timedelta(days=i)
        day_sales  = sale_qs.filter(created_at__date=d)
        revenue    = float(day_sales.aggregate(t=Sum('total_amount'))['t'] or 0)
        day_items  = SaleItem.objects.filter(sale__in=day_sales).select_related('product')
        cost       = sum(float(i.product.cost_price if i.product else 0) * i.quantity
                         for i in day_items)
        daily_data.append({'date': d.strftime('%d %b'), 'revenue': revenue,
                            'profit': round(revenue - cost, 2)})

    product_profit = {}
    for item in items:
        if item.product:
            name = item.product_name
            rev  = float(item.total_price)
            cost = float(item.product.cost_price) * item.quantity
            if name not in product_profit:
                product_profit[name] = {'revenue': 0, 'cost': 0, 'units': 0}
            product_profit[name]['revenue'] += rev
            product_profit[name]['cost']    += cost
            product_profit[name]['units']   += item.quantity

    top_profitable = sorted(
        [{'name': k, 'profit': round(v['revenue'] - v['cost'], 2),
          'units': v['units'], 'revenue': round(v['revenue'], 2)}
         for k, v in product_profit.items()],
        key=lambda x: x['profit'], reverse=True
    )[:10]

    return render(request, 'pos/profit.html', {
        'total_revenue':  total_revenue,
        'total_cost':     total_cost,
        'total_profit':   total_profit,
        'profit_margin':  profit_margin,
        'daily_data':     json.dumps(daily_data),
        'top_profitable': top_profitable,
        'period':         str(days),
    })


@store_required
def daily_summary_view(request):
    report_date_str = request.GET.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        report_date = date.fromisoformat(report_date_str)
    except ValueError:
        report_date = date.today()

    qs       = store_queryset(Sale, request).filter(
        created_at__date=report_date, status='completed'
    )
    items_qs = SaleItem.objects.filter(sale__in=qs).select_related('product')

    total_revenue = qs.aggregate(t=Sum('total_amount'))['t'] or 0
    total_cost    = sum(
        float(i.product.cost_price if i.product else 0) * i.quantity for i in items_qs
    )

    top_products = (
        SaleItem.objects.filter(sale__in=qs)
        .values('product_name')
        .annotate(units=Sum('quantity'), revenue=Sum('total_price'))
        .order_by('-revenue')[:10]
    )

    return render(request, 'pos/daily_summary.html', {
        'report_date':      report_date,
        'report_date_str':  report_date_str,
        'total_revenue':    total_revenue,
        'total_discount':   qs.aggregate(t=Sum('discount_amount'))['t'] or 0,
        'total_tax':        qs.aggregate(t=Sum('tax_amount'))['t'] or 0,
        'total_cash':       qs.filter(payment_method='cash').aggregate(t=Sum('total_amount'))['t'] or 0,
        'total_card':       qs.filter(payment_method='card').aggregate(t=Sum('total_amount'))['t'] or 0,
        'total_online':     qs.filter(payment_method='online').aggregate(t=Sum('total_amount'))['t'] or 0,
        'transaction_count':qs.count(),
        'top_products':     top_products,
        'net_profit':       float(total_revenue) - total_cost,
        'total_cost':       total_cost,
        'sales':            qs.order_by('created_at').select_related('cashier', 'customer'),
    })


@store_required
def daily_summary_pdf(request):
    report_date_str = request.GET.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        report_date = date.fromisoformat(report_date_str)
    except ValueError:
        report_date = date.today()

    qs       = store_queryset(Sale, request).filter(
        created_at__date=report_date, status='completed'
    )
    items_qs = SaleItem.objects.filter(sale__in=qs)
    top_products = (items_qs.values('product_name')
                    .annotate(units=Sum('quantity'), revenue=Sum('total_price'))
                    .order_by('-revenue')[:10])
    total_revenue = qs.aggregate(t=Sum('total_amount'))['t'] or 0
    total_cost    = sum(
        float(i.product.cost_price if i.product else 0) * i.quantity
        for i in items_qs.select_related('product')
    )
    net_profit = float(total_revenue) - total_cost

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, HRFlowable)
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT

        buf        = io.BytesIO()
        store_name = request.store.name if request.store else 'SmartPOS'
        doc        = SimpleDocTemplate(buf, pagesize=A4,
                                       topMargin=2*cm, bottomMargin=2*cm,
                                       leftMargin=2*cm, rightMargin=2*cm)

        title    = ParagraphStyle('title', fontSize=18, fontName='Helvetica-Bold',
                                  alignment=1, spaceAfter=6)
        subtitle = ParagraphStyle('sub', fontSize=11, alignment=1,
                                  textColor=colors.HexColor('#666666'), spaceAfter=4)
        heading  = ParagraphStyle('h', fontSize=12, fontName='Helvetica-Bold',
                                  spaceBefore=12, spaceAfter=6)
        hdr_fill = colors.HexColor('#1a1a2e')
        row_fills= [colors.HexColor('#f8f9fa'), colors.white]
        base_ts  = [
            ('BACKGROUND', (0,0), (-1,0), hdr_fill),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 10),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), row_fills),
            ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
            ('PADDING',    (0,0), (-1,-1), 8),
        ]

        story = [
            Paragraph(store_name, title),
            Paragraph('Daily Summary Report', subtitle),
            Paragraph(report_date.strftime('%A, %d %B %Y'), subtitle),
            Spacer(1, 0.4*cm),
            HRFlowable(width='100%', thickness=1, color=colors.HexColor('#333333')),
            Spacer(1, 0.3*cm),
        ]

        story.append(Paragraph('Sales Summary', heading))
        t1 = Table([
            ['Metric', 'Value'],
            ['Total Transactions', str(qs.count())],
            ['Gross Revenue',       f'Rs {total_revenue:.2f}'],
            ['Total Discounts',     f'- Rs {qs.aggregate(t=Sum("discount_amount"))["t"] or 0:.2f}'],
            ['Tax Collected',       f'Rs {qs.aggregate(t=Sum("tax_amount"))["t"] or 0:.2f}'],
            ['Estimated Profit',    f'Rs {net_profit:.2f}'],
        ], colWidths=[10*cm, 7*cm])
        t1.setStyle(TableStyle(base_ts + [('FONTNAME', (0,5), (-1,5), 'Helvetica-Bold')]))
        story.append(t1)

        story.append(Paragraph('Payment Breakdown', heading))
        t2 = Table([
            ['Method', 'Amount'],
            ['Cash',   f'Rs {qs.filter(payment_method="cash").aggregate(t=Sum("total_amount"))["t"] or 0:.2f}'],
            ['Card',   f'Rs {qs.filter(payment_method="card").aggregate(t=Sum("total_amount"))["t"] or 0:.2f}'],
            ['Online', f'Rs {qs.filter(payment_method="online").aggregate(t=Sum("total_amount"))["t"] or 0:.2f}'],
            ['Total',  f'Rs {total_revenue:.2f}'],
        ], colWidths=[10*cm, 7*cm])
        t2.setStyle(TableStyle(base_ts + [('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold')]))
        story.append(t2)

        if top_products:
            story.append(Paragraph('Top Selling Products', heading))
            rows = [['#', 'Product', 'Units', 'Revenue']]
            for i, p in enumerate(top_products, 1):
                rows.append([str(i), p['product_name'], str(p['units']), f"Rs {p['revenue']:.2f}"])
            t3 = Table(rows, colWidths=[1*cm, 10*cm, 3*cm, 4*cm])
            t3.setStyle(TableStyle(base_ts))
            story.append(t3)

        if qs.exists():
            story.append(Paragraph('Transaction Details', heading))
            rows = [['Sale #', 'Time', 'Cashier', 'Customer', 'Payment', 'Total']]
            for s in qs.order_by('created_at'):
                rows.append([
                    s.sale_number, s.created_at.strftime('%H:%M'),
                    s.cashier.username if s.cashier else '-',
                    s.customer.name if s.customer else 'Walk-in',
                    s.get_payment_method_display(),
                    f'Rs {s.total_amount:.2f}',
                ])
            t4 = Table(rows, colWidths=[3*cm, 2*cm, 3*cm, 4*cm, 2.5*cm, 3.5*cm])
            t4.setStyle(TableStyle(base_ts + [('FONTSIZE', (0,0), (-1,-1), 8)]))
            story.append(t4)

        story += [
            Spacer(1, 0.5*cm),
            HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#cccccc')),
            Paragraph(
                f"Generated on {date.today().strftime('%d %b %Y')} | {store_name}",
                ParagraphStyle('foot', fontSize=8,
                               textColor=colors.HexColor('#999999'),
                               alignment=1, spaceBefore=6)
            ),
        ]

        doc.build(story)
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/pdf')
        response['Content-Disposition'] = \
            f'attachment; filename="daily_report_{report_date_str}.pdf"'
        return response

    except ImportError:
        messages.error(request, 'reportlab is required. Run: pip install reportlab')
        return redirect('sales')


@store_required
def loyalty_report(request):
    customers = store_queryset(Customer, request).filter(
        loyalty_points__gt=0
    ).order_by('-loyalty_points')

    sale_qs = store_queryset(Sale, request).filter(status='completed')

    return render(request, 'pos/loyalty.html', {
        'customers':           customers,
        'total_points_issued': customers.aggregate(t=Sum('loyalty_points'))['t'] or 0,
        'total_points_used':   sale_qs.aggregate(t=Sum('loyalty_points_used'))['t'] or 0,
        'total_points_earned': sale_qs.aggregate(t=Sum('loyalty_points_earned'))['t'] or 0,
        'points_value':        float(
            customers.aggregate(t=Sum('loyalty_points'))['t'] or 0
        ) * 0.5,
    })


# ══════════════════════════════════════════════════════════════════════════════
# EXPENSES
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def expenses_list(request):
    store = request.store
    qs    = store_queryset(Expense, request).select_related(
        'category', 'added_by'
    ).order_by('-date', '-created_at')

    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    total      = qs.aggregate(t=Sum('amount'))['t'] or 0
    categories = store_queryset(ExpenseCategory, request)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_expense':
            cat_id = request.POST.get('category')
            try:
                amount = Decimal(request.POST.get('amount', ''))
                if amount <= 0:
                    raise ValueError
            except (InvalidOperation, ValueError):
                messages.error(request, 'Invalid expense amount.')
                return redirect('expenses')

            title = request.POST.get('title', '').strip()
            if not title:
                messages.error(request, 'Expense title is required.')
                return redirect('expenses')

            exp = Expense.objects.create(
                store=store,
                category=get_object_or_404(ExpenseCategory, id=cat_id, store=store) if cat_id else None,
                title=title,
                amount=amount,
                date=request.POST.get('date', date.today()),
                notes=request.POST.get('notes', ''),
                added_by=request.user,
            )
            audit(request, 'expense_add',
                  f'Expense added: "{exp.title}" — Rs {exp.amount}')
            _trigger_gdrive_auto_backup(store, 'expense', 'gdrive_backup_on_expense')
            messages.success(request, 'Expense recorded.')
            return redirect('expenses')

        elif action == 'add_category':
            cat_name = request.POST.get('cat_name', '').strip()
            if not cat_name:
                messages.error(request, 'Category name is required.')
                return redirect('expenses')
            ExpenseCategory.objects.create(name=cat_name, store=store)
            messages.success(request, 'Category added.')
            return redirect('expenses')

        elif action == 'delete_expense':
            exp_id = request.POST.get('expense_id')
            exp_obj = store_queryset(Expense, request).filter(id=exp_id).first()
            if exp_obj:
                audit(request, 'expense_delete',
                      f'Expense deleted: "{exp_obj.title}" — Rs {exp_obj.amount}')
                exp_obj.delete()
            messages.success(request, 'Expense deleted.')
            return redirect('expenses')

    return render(request, 'pos/expenses.html', {
        'expenses':   qs,
        'categories': categories,
        'total':      total,
        'date_from':  date_from,
        'date_to':    date_to,
    })


# ══════════════════════════════════════════════════════════════════════════════
# SHIFT / CASH DRAWER
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def shift_view(request):
    store      = request.store
    open_shift = Shift.objects.filter(cashier=request.user, status='open').first()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'open_shift' and not open_shift:
            try:
                opening_cash = Decimal(request.POST.get('opening_cash', 0))
            except InvalidOperation:
                opening_cash = Decimal('0')
            s = Shift.objects.create(
                cashier=request.user, store=store,
                opening_cash=opening_cash, status='open',
            )
            audit(request, 'shift_open',
                  f'Shift opened — Opening cash: Rs {s.opening_cash}')
            messages.success(request, 'Shift opened successfully.')
            return redirect('shift')

        elif action == 'close_shift' and open_shift:
            try:
                closing_cash = Decimal(request.POST.get('closing_cash', 0))
            except InvalidOperation:
                closing_cash = Decimal('0')

            shift_sales = store_queryset(Sale, request).filter(
                cashier=request.user,
                created_at__gte=open_shift.opened_at,
                status='completed', payment_method='cash',
            )
            expected   = open_shift.opening_cash + (
                shift_sales.aggregate(t=Sum('total_amount'))['t'] or Decimal('0')
            )
            difference = closing_cash - expected
            all_sales  = store_queryset(Sale, request).filter(
                cashier=request.user,
                created_at__gte=open_shift.opened_at,
                status='completed',
            )
            open_shift.closing_cash       = closing_cash
            open_shift.expected_cash      = expected
            open_shift.cash_difference    = difference
            open_shift.total_sales        = all_sales.aggregate(t=Sum('total_amount'))['t'] or 0
            open_shift.total_transactions = all_sales.count()
            open_shift.notes              = request.POST.get('notes', '')
            open_shift.status             = 'closed'
            open_shift.closed_at          = timezone.now()
            open_shift.save()
            audit(request, 'shift_close',
                  f'Shift closed — Closing: Rs {closing_cash} — '
                  f'Expected: Rs {expected} — Diff: Rs {difference}')
            messages.success(request, 'Shift closed successfully.')
            return redirect('shift')

    shift_stats = None
    if open_shift:
        shift_sales = store_queryset(Sale, request).filter(
            cashier=request.user,
            created_at__gte=open_shift.opened_at,
            status='completed',
        )
        shift_stats = {
            'total':  shift_sales.aggregate(t=Sum('total_amount'))['t'] or 0,
            'count':  shift_sales.count(),
            'cash':   shift_sales.filter(payment_method='cash').aggregate(t=Sum('total_amount'))['t'] or 0,
            'card':   shift_sales.filter(payment_method='card').aggregate(t=Sum('total_amount'))['t'] or 0,
            'online': shift_sales.filter(payment_method='online').aggregate(t=Sum('total_amount'))['t'] or 0,
        }

    return render(request, 'pos/shift.html', {
        'open_shift':    open_shift,
        'shift_stats':   shift_stats,
        'recent_shifts': Shift.objects.filter(
            cashier=request.user
        ).order_by('-opened_at')[:10],
    })


# ══════════════════════════════════════════════════════════════════════════════
# REORDER ALERTS
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def reorder_alerts(request):
    qs = store_queryset(Product, request).filter(
        is_active=True,
        stock_quantity__lte=models.F('low_stock_threshold')
    ).select_related('supplier', 'category').order_by('stock_quantity')

    return render(request, 'pos/reorder.html', {
        'out_of_stock': qs.filter(stock_quantity=0),
        'low_stock':    qs.filter(stock_quantity__gt=0),
        'total_alerts': qs.count(),
    })


# ══════════════════════════════════════════════════════════════════════════════
# BARCODE LABELS
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def barcode_labels(request):
    products     = store_queryset(Product, request).filter(is_active=True)
    selected_ids = request.GET.getlist('products')
    selected_products = (
        store_queryset(Product, request).filter(id__in=selected_ids)
        if selected_ids else []
    )
    return render(request, 'pos/barcode_labels.html', {
        'products':          products,
        'selected_products': selected_products,
        'copies':            max(1, min(int(request.GET.get('copies', 1)), 100)),
        'selected_ids':      selected_ids,
    })


# ══════════════════════════════════════════════════════════════════════════════
# STORE SETTINGS
# ══════════════════════════════════════════════════════════════════════════════

@store_required
@role_required('admin', 'manager')
def store_settings(request):
    store        = request.store
    settings_obj, _ = StoreSettings.objects.get_or_create(
        store=store,
        defaults={'receipt_footer': 'Thank you for shopping with us!'}
    )

    if request.method == 'POST':
        store.name    = request.POST.get('store_name', store.name).strip()
        store.phone   = request.POST.get('phone', store.phone)
        store.email   = request.POST.get('email', store.email)
        store.address = request.POST.get('address', store.address)
        try:
            store.tax_rate = Decimal(request.POST.get('tax_rate', store.tax_rate))
        except InvalidOperation:
            pass
        store.save()

        settings_obj.receipt_header           = request.POST.get('receipt_header', '')
        settings_obj.receipt_footer           = request.POST.get('receipt_footer',
                                                'Thank you for shopping with us!')
        settings_obj.show_cashier_on_receipt  = 'show_cashier' in request.POST
        settings_obj.show_customer_on_receipt = 'show_customer' in request.POST
        settings_obj.currency_symbol          = request.POST.get('currency_symbol', 'Rs')

        if 'logo' in request.FILES:
            logo_file = request.FILES['logo']
            err = _validate_image(logo_file)
            if err:
                messages.error(request, err)
                return redirect('store_settings')
            settings_obj.logo = logo_file

        settings_obj.save()
        audit(request, 'settings_change',
              f'Store settings updated by {request.user.username}')
        messages.success(request, 'Settings saved successfully.')
        return redirect('store_settings')

    return render(request, 'pos/store_settings.html', {
        'store': store, 'settings': settings_obj
    })


# ══════════════════════════════════════════════════════════════════════════════
# USER MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

@store_required
@role_required('admin', 'manager')
def user_list(request):
    users = User.objects.filter(
        profile__store=request.store
    ).select_related('profile')
    return render(request, 'pos/users.html', {'users': users})


@store_required
@role_required('admin', 'manager')
def user_add(request):
    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        password   = request.POST.get('password', '')
        first_name = request.POST.get('first_name', '')
        last_name  = request.POST.get('last_name', '')
        role       = request.POST.get('role', 'cashier')

        # FIX SEC-PASSWD: Enforce minimum 8-char password for staff-created users too
        if len(password) < 8:
            messages.error(request, 'Password must be at least 8 characters.')
            return render(request, 'pos/user_form.html', {'action': 'Add'})

        # FIX SEC-RBAC: Managers cannot create admin accounts
        if role not in ('cashier', 'manager', 'admin'):
            role = 'cashier'
        try:
            if request.user.profile.role == 'manager' and role == 'admin':
                messages.error(request, 'Managers cannot create admin accounts.')
                return render(request, 'pos/user_form.html', {'action': 'Add'})
        except Exception:
            pass

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
        else:
            user = User.objects.create_user(
                username=username, password=password,
                first_name=first_name, last_name=last_name
            )
            UserProfile.objects.create(user=user, role=role, store=request.store)
            audit(request, 'user_add', f'User "{username}" created with role "{role}"')
            messages.success(request, f'User {username} created.')
            return redirect('users')
    return render(request, 'pos/user_form.html', {'action': 'Add'})


# ══════════════════════════════════════════════════════════════════════════════
# STORES  (superadmin / admin only)
# ══════════════════════════════════════════════════════════════════════════════

@store_required
@role_required('admin')
def store_list(request):
    stores = Store.objects.all() if request.user.is_superuser else Store.objects.filter(
        id=request.store.id
    )
    return render(request, 'pos/stores.html', {'stores': stores})


@store_required
@role_required('admin')
def store_add(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Store name is required.')
            return redirect('store_add')
        Store.objects.create(
            name=name,
            address=request.POST.get('address', ''),
            phone=request.POST.get('phone', ''),
            email=request.POST.get('email', ''),
            tax_rate=request.POST.get('tax_rate', 0),
        )
        messages.success(request, 'Store added.')
        return redirect('stores')
    return render(request, 'pos/store_form.html', {'action': 'Add'})


# ══════════════════════════════════════════════════════════════════════════════
# PIN MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

@store_required
@role_required('admin', 'manager')
def manage_pins(request):
    users = User.objects.filter(
        profile__store=request.store
    ).select_related('profile')

    if request.method == 'POST':
        action  = request.POST.get('action')
        user_id = request.POST.get('user_id')
        try:
            target_user = User.objects.get(id=user_id, profile__store=request.store)
            profile     = target_user.profile
            if action == 'set_pin':
                pin = request.POST.get('pin', '').strip()
                if len(pin) != 4 or not pin.isdigit():
                    messages.error(request, 'PIN must be exactly 4 digits.')
                else:
                    # FIX SEC-PIN: Check hash uniqueness, not plaintext
                    # We must check all profiles to see if any other user has this pin
                    pin_conflict = False
                    for p in UserProfile.objects.exclude(user=target_user):
                        if p.check_pin(pin):
                            pin_conflict = True
                            break
                    if pin_conflict:
                        messages.error(request, 'PIN already used by another user.')
                    else:
                        profile.set_pin(pin)
                        profile.save()
                        messages.success(request, f'PIN set for {target_user.username}.')
            elif action == 'clear_pin':
                profile.clear_pin()
                profile.save()
                messages.success(request, f'PIN cleared for {target_user.username}.')
        except User.DoesNotExist:
            messages.error(request, 'User not found.')
        return redirect('manage_pins')

    return render(request, 'pos/manage_pins.html', {'users': users})


@store_required
@role_required('admin', 'manager')
def reset_cashier(request):
    users = User.objects.filter(
        profile__store=request.store
    ).select_related('profile')

    if request.method == 'POST':
        action  = request.POST.get('action')
        user_id = request.POST.get('user_id')
        try:
            target_user = User.objects.get(id=user_id, profile__store=request.store)
            if action == 'reset_password':
                new_password = request.POST.get('new_password', '').strip()
                if len(new_password) < 8:
                    messages.error(request, 'Password must be at least 8 characters.')
                else:
                    target_user.set_password(new_password)
                    target_user.save()
                    messages.success(request, f'Password reset for {target_user.username}.')
            elif action == 'reset_pin':
                target_user.profile.clear_pin()
                target_user.profile.save()
                messages.success(request, f'PIN cleared for {target_user.username}.')
            elif action == 'toggle_active':
                target_user.is_active = not target_user.is_active
                target_user.save()
                status = 'activated' if target_user.is_active else 'deactivated'
                messages.success(request, f'User {target_user.username} {status}.')
        except User.DoesNotExist:
            messages.error(request, 'User not found.')
        return redirect('reset_cashier')

    return render(request, 'pos/reset_cashier.html', {'users': users})


# ══════════════════════════════════════════════════════════════════════════════
# AUDIT LOG
# ══════════════════════════════════════════════════════════════════════════════

@store_required
@role_required('admin', 'manager')
def audit_log_view(request):
    qs = store_queryset(AuditLog, request).select_related('user').order_by('-timestamp')

    action_filter = request.GET.get('action', '')
    user_filter   = request.GET.get('user', '')
    date_from     = request.GET.get('date_from', '')
    date_to       = request.GET.get('date_to', '')

    if action_filter:
        qs = qs.filter(action=action_filter)
    if user_filter:
        qs = qs.filter(user__username__icontains=user_filter)
    if date_from:
        qs = qs.filter(timestamp__date__gte=date_from)
    if date_to:
        qs = qs.filter(timestamp__date__lte=date_to)

    return render(request, 'pos/audit_log.html', {
        'logs':           qs[:500],
        'action_choices': AuditLog.ACTION_CHOICES,
        'action_filter':  action_filter,
        'user_filter':    user_filter,
        'date_from':      date_from,
        'date_to':        date_to,
        'total':          qs.count(),
    })


# ══════════════════════════════════════════════════════════════════════════════
# TAX ENGINE
# ══════════════════════════════════════════════════════════════════════════════

@store_required
@role_required('admin', 'manager')
def tax_rules_view(request):
    store      = request.store
    rules      = store_queryset(TaxRule, request)
    categories = store_queryset(Category, request)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            cat_id = request.POST.get('category')
            name   = request.POST.get('name', '').strip()
            if not name:
                messages.error(request, 'Tax rule name is required.')
                return redirect('tax_rules')
            try:
                rate = Decimal(request.POST.get('rate', ''))
                if rate < 0 or rate > 100:
                    raise ValueError
            except (InvalidOperation, ValueError):
                messages.error(request, 'Rate must be between 0 and 100.')
                return redirect('tax_rules')

            tax_type = request.POST.get('tax_type', 'SALES_TAX')
            tax_mode = request.POST.get('tax_mode', 'exclusive')
            apply_to = request.POST.get('apply_to', 'all')

            if tax_type not in dict(TaxRule.TAX_TYPE_CHOICES):
                tax_type = 'SALES_TAX'
            if tax_mode not in ('inclusive', 'exclusive'):
                tax_mode = 'exclusive'
            if apply_to not in ('all', 'category'):
                apply_to = 'all'

            TaxRule.objects.create(
                store=store, name=name, tax_type=tax_type, rate=rate,
                tax_mode=tax_mode, apply_to=apply_to,
                category=get_object_or_404(Category, id=cat_id) if cat_id else None,
                is_active=True,
            )
            audit(request, 'settings_change',
                  f'Tax rule added: {name} @ {rate}%')
            messages.success(request, 'Tax rule added.')
            return redirect('tax_rules')

        elif action == 'toggle':
            rule = get_object_or_404(TaxRule, id=request.POST.get('rule_id'), store=store)
            rule.is_active = not rule.is_active
            rule.save()
            status = 'enabled' if rule.is_active else 'disabled'
            audit(request, 'settings_change', f'Tax rule "{rule.name}" {status}')
            messages.success(request, f'Tax rule {status}.')
            return redirect('tax_rules')

        elif action == 'delete':
            rule = get_object_or_404(TaxRule, id=request.POST.get('rule_id'), store=store)
            audit(request, 'settings_change', f'Tax rule deleted: {rule.name}')
            rule.delete()
            messages.success(request, 'Tax rule deleted.')
            return redirect('tax_rules')

    return render(request, 'pos/tax_rules.html', {
        'rules':            rules,
        'categories':       categories,
        'tax_type_choices': TaxRule.TAX_TYPE_CHOICES,
        'apply_choices':    TaxRule.APPLY_CHOICES,
        'apply_to_choices': TaxRule.APPLY_TO_CHOICES,
        'store_tax_rate':   store.tax_rate,
    })


# ══════════════════════════════════════════════════════════════════════════════
# CURRENCY
# ══════════════════════════════════════════════════════════════════════════════

COMMON_CURRENCIES = [
    ('PKR', 'Rs',  'Pakistani Rupee'),
    ('USD', '$',   'US Dollar'),
    ('EUR', '€',   'Euro'),
    ('GBP', '£',   'British Pound'),
    ('AED', 'AED', 'UAE Dirham'),
    ('SAR', 'SAR', 'Saudi Riyal'),
    ('INR', '₹',   'Indian Rupee'),
    ('BDT', '৳',   'Bangladeshi Taka'),
    ('MYR', 'RM',  'Malaysian Ringgit'),
    ('TRY', '₺',   'Turkish Lira'),
]

_VALID_CURRENCY_CODES = {c[0] for c in COMMON_CURRENCIES}


@store_required
@role_required('admin', 'manager')
def currency_settings_view(request):
    store        = request.store
    settings_obj, _ = StoreSettings.objects.get_or_create(store=store)

    if request.method == 'POST':
        symbol = request.POST.get('custom_symbol', '').strip() or request.POST.get('currency_symbol', 'Rs')
        code   = request.POST.get('currency_code', 'PKR').upper()
        # Only allow known codes to prevent arbitrary data
        if code not in _VALID_CURRENCY_CODES:
            code = 'PKR'
        try:
            rate = Decimal(request.POST.get('exchange_rate', '1'))
            if rate <= 0:
                rate = Decimal('1')
        except InvalidOperation:
            rate = Decimal('1')

        settings_obj.currency_code   = code
        settings_obj.currency_symbol = symbol
        settings_obj.exchange_rate   = rate
        settings_obj.save()
        audit(request, 'settings_change',
              f'Currency changed to {code} ({symbol})')
        messages.success(request, f'Currency updated to {symbol}.')
        return redirect('currency_settings')

    return render(request, 'pos/currency_settings.html', {
        'settings':   settings_obj,
        'currencies': COMMON_CURRENCIES,
    })


@store_required
@role_required('admin', 'manager')
def fetch_live_rate(request):
    # FIX SEC-SSRF: Validate currency codes are alpha-only before inserting in URL
    target  = _re.sub(r'[^A-Z]', '', request.GET.get('target', 'USD').upper())[:3]
    base    = _re.sub(r'[^A-Z]', '', request.GET.get('base', 'PKR').upper())[:3]
    do_save = request.GET.get('save', '0') == '1'

    if not target or not base:
        return JsonResponse({'success': False, 'error': 'Invalid currency code'})

    store        = request.store
    settings_obj, _ = StoreSettings.objects.get_or_create(store=store)
    api_key      = settings_obj.exchange_rate_api_key.strip()

    try:
        import urllib.request as urlreq
        if api_key:
            url    = f'https://v6.exchangerate-api.com/v6/{api_key}/latest/{base}'
            source = 'exchangerate-api.com (API key)'
        else:
            url    = f'https://open.er-api.com/v6/latest/{base}'
            source = 'open.er-api.com (free)'

        with urlreq.urlopen(url, timeout=8) as resp:  # noqa: S310
            data = json.loads(resp.read().decode())

        if data.get('result') != 'success':
            return JsonResponse({'success': False,
                                 'error': data.get('error-type', 'API error')})

        rates = data.get('rates', {})
        if target not in rates:
            return JsonResponse({'success': False, 'error': f'{target} not found'})

        rate = Decimal(str(rates[target])).quantize(Decimal('0.000001'))
        if do_save:
            settings_obj.exchange_rate = rate
            settings_obj.save()
            audit(request, 'settings_change',
                  f'Live rate saved: 1 {base} = {rate} {target}')

        return JsonResponse({
            'success':      True,
            'rate':         str(rate),
            'base':         base,
            'target':       target,
            'last_updated': data.get('time_last_update_utc', 'Unknown'),
            'source':       source,
            'used_api_key': bool(api_key),
        })
    except Exception as e:
        logger.error(f"fetch_live_rate error: {e}")
        return JsonResponse({'success': False, 'error': 'Rate fetch failed'})


@store_required
@role_required('admin', 'manager')
@require_POST
def save_api_key(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})

    key  = data.get('api_key', '').strip()
    s, _ = StoreSettings.objects.get_or_create(store=request.store)
    s.exchange_rate_api_key = key
    s.save()
    audit(request, 'settings_change',
          f'Exchange rate API key {"saved" if key else "removed"}')
    return JsonResponse({'success': True})


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def export_sales_csv(request):
    qs = store_queryset(Sale, request).filter(status='completed').order_by('-created_at')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sales_report.csv"'
    writer = csv.writer(response)
    writer.writerow(['Sale #', 'Date', 'Cashier', 'Customer', 'Items',
                     'Subtotal', 'Discount', 'Tax', 'Total', 'Payment', 'Points Earned'])
    for s in qs:
        writer.writerow([
            s.sale_number, s.created_at.strftime('%Y-%m-%d %H:%M'),
            s.cashier.username if s.cashier else '',
            s.customer.name if s.customer else '',
            s.items.count(),
            s.subtotal, s.discount_amount, s.tax_amount, s.total_amount,
            s.get_payment_method_display(), s.loyalty_points_earned,
        ])
    return response


@store_required
def export_products_csv(request):
    qs = store_queryset(Product, request).filter(is_active=True)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'
    writer = csv.writer(response)
    writer.writerow(['Name', 'Barcode', 'Category', 'Price', 'Cost Price', 'Stock', 'Supplier'])
    for p in qs:
        writer.writerow([
            p.name, p.barcode,
            p.category.name if p.category else '',
            p.price, p.cost_price, p.stock_quantity,
            p.supplier.name if p.supplier else '',
        ])
    return response


# ══════════════════════════════════════════════════════════════════════════════
# BACKUP & RESTORE
# ══════════════════════════════════════════════════════════════════════════════

@store_required
@role_required('admin', 'manager')
def backup_database(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        store     = request.store
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename  = f"smartpos_backup_{timestamp}.xlsx"
        wb        = openpyxl.Workbook()

        header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        header_fill  = PatternFill("solid", start_color="1a1a2e")
        header_align = Alignment(horizontal="center", vertical="center")
        row_fill_a   = PatternFill("solid", start_color="F8F9FA")
        row_fill_b   = PatternFill("solid", start_color="FFFFFF")

        def style_sheet(ws, headers, rows):
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font; cell.fill = header_fill
                cell.alignment = header_align
            ws.row_dimensions[1].height = 20
            for i, row in enumerate(rows, start=2):
                ws.append(row)
                fill = row_fill_a if i % 2 == 0 else row_fill_b
                for cell in ws[i]:
                    cell.fill = fill
                    cell.font = Font(name="Arial", size=9)
                    cell.alignment = Alignment(vertical="center")
            for col_idx, _ in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    (len(str(ws.cell(r, col_idx).value or ""))
                     for r in range(1, ws.max_row + 1)), default=10
                )
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        ws1 = wb.active; ws1.title = "Sales"
        style_sheet(ws1,
            ["Sale #","Date","Time","Cashier","Customer","Subtotal (Rs)","Discount (Rs)",
             "Tax (Rs)","Total (Rs)","Payment Method","Cash Received","Change",
             "Points Earned","Points Used","Status"],
            [[s.sale_number, s.created_at.strftime("%d/%m/%Y"), s.created_at.strftime("%H:%M"),
              s.cashier.username if s.cashier else "-",
              s.customer.name if s.customer else "Walk-in",
              float(s.subtotal), float(s.discount_amount), float(s.tax_amount),
              float(s.total_amount), s.get_payment_method_display(),
              float(s.amount_received), float(s.change_amount),
              s.loyalty_points_earned, s.loyalty_points_used, s.status.title()]
             for s in store_queryset(Sale, request).order_by("-created_at")]
        )

        ws2 = wb.create_sheet("Sale Items")
        items_qs = SaleItem.objects.filter(
            sale__store=store
        ).select_related("sale").order_by("-sale__created_at")
        style_sheet(ws2,
            ["Sale #","Date","Product Name","Barcode","Qty","Unit Price (Rs)","Total (Rs)","Returned Qty"],
            [[i.sale.sale_number, i.sale.created_at.strftime("%d/%m/%Y"),
              i.product_name, i.product_barcode, i.quantity,
              float(i.unit_price), float(i.total_price), i.returned_quantity]
             for i in items_qs]
        )

        ws3 = wb.create_sheet("Products")
        style_sheet(ws3,
            ["Name","Barcode","Category","Selling Price (Rs)","Cost Price (Rs)",
             "Stock Qty","Low Stock Threshold","Supplier","Status"],
            [[p.name, p.barcode, p.category.name if p.category else "-",
              float(p.price), float(p.cost_price), p.stock_quantity,
              p.low_stock_threshold, p.supplier.name if p.supplier else "-",
              p.stock_status.title()]
             for p in store_queryset(Product, request).filter(is_active=True)]
        )

        ws4 = wb.create_sheet("Customers")
        style_sheet(ws4,
            ["Name","Phone","Email","Address","Loyalty Points","Total Purchases","Total Spent (Rs)","Joined"],
            [[c.name, c.phone, c.email, c.address, c.loyalty_points,
              c.total_purchases(), float(c.total_spent()), c.created_at.strftime("%d/%m/%Y")]
             for c in store_queryset(Customer, request)]
        )

        ws5 = wb.create_sheet("Expenses")
        style_sheet(ws5,
            ["Date","Title","Category","Amount (Rs)","Added By","Notes"],
            [[e.date.strftime("%d/%m/%Y"), e.title,
              e.category.name if e.category else "-", float(e.amount),
              e.added_by.username if e.added_by else "-", e.notes]
             for e in store_queryset(Expense, request).order_by("-date")]
        )

        ws6 = wb.create_sheet("Suppliers")
        style_sheet(ws6,
            ["Name","Phone","Email","Address","Balance Owed (Rs)"],
            [[s.name, s.phone, s.email, s.address, float(s.balance_owed)]
             for s in store_queryset(Supplier, request)]
        )

        ws7 = wb.create_sheet("Summary", 0); ws7.title = "Summary"
        ws7.column_dimensions["A"].width = 32
        ws7.column_dimensions["B"].width = 22
        total_revenue  = store_queryset(Sale, request).filter(
            status="completed").aggregate(t=Sum("total_amount"))["t"] or 0
        total_expenses = store_queryset(Expense, request).aggregate(t=Sum("amount"))["t"] or 0
        total_supplier = store_queryset(Supplier, request).aggregate(t=Sum("balance_owed"))["t"] or 0
        ws7["A1"] = f"SmartPOS — {store.name} — Backup Summary"
        ws7["A1"].font = Font(bold=True, name="Arial", size=14, color="1a1a2e")
        ws7["A2"] = f"Generated: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        ws7["A2"].font = Font(name="Arial", size=9, color="888888")
        for row_num, label, value in [
            (4,  "SALES & REVENUE", ""),
            (5,  "Total Completed Sales",
             store_queryset(Sale, request).filter(status="completed").count()),
            (6,  "Total Revenue (Rs)", float(total_revenue)),
            (7,  "Total Expenses (Rs)", float(total_expenses)),
            (8,  "Net Profit Estimate (Rs)", float(total_revenue) - float(total_expenses)),
            (9,  "Supplier Balance Owed (Rs)", float(total_supplier)),
            (11, "INVENTORY", ""),
            (12, "Total Active Products",
             store_queryset(Product, request).filter(is_active=True).count()),
            (13, "Out of Stock",
             store_queryset(Product, request).filter(is_active=True, stock_quantity=0).count()),
            (16, "CUSTOMERS", ""),
            (17, "Total Customers", store_queryset(Customer, request).count()),
        ]:
            ca = ws7.cell(row=row_num, column=1, value=label)
            cb = ws7.cell(row=row_num, column=2, value=value)
            if value == "":
                ca.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
                ca.fill = cb.fill = PatternFill("solid", start_color="1a1a2e")
            else:
                ca.font = Font(bold=True, name="Arial", size=10)
                fill = row_fill_a if row_num % 2 == 0 else row_fill_b
                ca.fill = cb.fill = fill

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)

        BackupLog.objects.create(created_by=request.user, filename=filename, store=store)
        audit(request, 'backup', f'Backup created: {filename}')

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        logger.error(f"backup_database error: {e}")
        messages.error(request, 'Backup failed. Please try again.')
        return redirect('dashboard')


@store_required
@role_required('admin')
def restore_backup(request):
    results = []
    errors  = []
    store   = request.store

    if request.method == 'POST' and request.FILES.get('backup_file'):
        backup_file = request.FILES['backup_file']

        # FIX SEC-UPLOAD: Validate backup file type and size
        if backup_file.size > 20 * 1024 * 1024:  # 20 MB limit
            messages.error(request, 'Backup file too large (max 20 MB).')
            return redirect('restore_backup')

        allowed_xlsx_types = {
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
        }
        if backup_file.content_type not in allowed_xlsx_types:
            messages.error(request, 'Only Excel (.xlsx) backup files are accepted.')
            return redirect('restore_backup')

        import openpyxl
        try:
            wb = openpyxl.load_workbook(backup_file)

            if 'Products' in wb.sheetnames:
                count = 0
                for row in wb['Products'].iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    try:
                        name, barcode, cat_name, price, cost_price, stock, threshold, sup_name, status = (list(row) + [None]*9)[:9]
                        if not barcode:
                            continue
                        cat = Category.objects.filter(name=cat_name, store=store).first() if cat_name and cat_name != '-' else None
                        sup = Supplier.objects.filter(name=sup_name, store=store).first() if sup_name and sup_name != '-' else None
                        Product.objects.update_or_create(
                            barcode=str(barcode), store=store,
                            defaults={
                                'name': name or 'Unknown',
                                'price': float(price or 0),
                                'cost_price': float(cost_price or 0),
                                'stock_quantity': int(stock or 0),
                                'low_stock_threshold': int(threshold or 10),
                                'category': cat, 'supplier': sup,
                                'is_active': True,
                            }
                        )
                        count += 1
                    except Exception as e:
                        errors.append(f'Product row: {e}')
                results.append(f'✅ Products: {count} restored')

            if 'Customers' in wb.sheetnames:
                count = 0
                for row in wb['Customers'].iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    try:
                        name, phone, email, address, points = (list(row) + [None]*5)[:5]
                        Customer.objects.get_or_create(
                            name=str(name), phone=str(phone or ''),
                            store=store,
                            defaults={'email': str(email or ''), 'address': str(address or ''),
                                      'loyalty_points': int(points or 0)}
                        )
                        count += 1
                    except Exception as e:
                        errors.append(f'Customer row: {e}')
                results.append(f'✅ Customers: {count} restored')

            if 'Expenses' in wb.sheetnames:
                count = 0
                for row in wb['Expenses'].iter_rows(min_row=2, values_only=True):
                    if not row[1]:
                        continue
                    try:
                        exp_date, title, cat_name, amount, added_by, notes = (list(row) + [None]*6)[:6]
                        cat = None
                        if cat_name and cat_name != '-':
                            cat, _ = ExpenseCategory.objects.get_or_create(
                                name=str(cat_name), store=store)
                        if isinstance(exp_date, str):
                            from datetime import datetime
                            exp_date = datetime.strptime(exp_date, '%d/%m/%Y').date()
                        Expense.objects.get_or_create(
                            title=str(title), date=exp_date,
                            amount=float(amount or 0), store=store,
                            defaults={'category': cat, 'notes': str(notes or ''),
                                      'added_by': request.user}
                        )
                        count += 1
                    except Exception as e:
                        errors.append(f'Expense row: {e}')
                results.append(f'✅ Expenses: {count} restored')

            if results:
                messages.success(request, f'Restore complete: {", ".join(results)}')
                audit(request, 'restore', f'Backup restored — {", ".join(results)}')
            if errors:
                messages.warning(request, f'{len(errors)} rows had errors (skipped).')

        except Exception as e:
            logger.error(f"restore_backup error: {e}")
            messages.error(request, 'Restore failed. File may be corrupted.')
        return redirect('restore_backup')

    return render(request, 'pos/restore_backup.html', {
        'results':     results,
        'errors':      errors,
        'backup_logs': BackupLog.objects.filter(store=store).order_by('-created_at')[:10],
    })


# ══════════════════════════════════════════════════════════════════════════════
# PURCHASE INVOICES
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def purchase_invoice(request, purchase_id):
    purchase     = get_object_or_404(StockPurchase, id=purchase_id, store=request.store)
    settings_obj, _ = StoreSettings.objects.get_or_create(store=request.store)
    return render(request, 'pos/purchase_invoice.html', {
        'purchase': purchase, 'settings': settings_obj, 'store': request.store,
    })


@store_required
def purchase_invoice_pdf(request, purchase_id):
    purchase = get_object_or_404(StockPurchase, id=purchase_id, store=request.store)
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, HRFlowable)
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT

        buf  = io.BytesIO()
        doc  = SimpleDocTemplate(buf, pagesize=A4,
                                 topMargin=2*cm, bottomMargin=2*cm,
                                 leftMargin=2*cm, rightMargin=2*cm)
        store_name = request.store.name
        center = ParagraphStyle('c', alignment=TA_CENTER, fontSize=10)
        right  = ParagraphStyle('r', alignment=TA_RIGHT,  fontSize=10)
        normal = ParagraphStyle('n', fontSize=10, spaceAfter=4)
        title_s= ParagraphStyle('t', alignment=TA_CENTER, fontSize=16,
                                fontName='Helvetica-Bold', spaceAfter=4)
        heading= ParagraphStyle('h', fontSize=12, fontName='Helvetica-Bold',
                                spaceBefore=12, spaceAfter=6)

        story = [
            Paragraph(store_name, title_s),
            Paragraph('Stock Purchase Invoice', center),
            Spacer(1, 0.3*cm),
            HRFlowable(width='100%', thickness=1, color=colors.black),
            Spacer(1, 0.3*cm),
        ]

        info_table = Table([
            ['Invoice #:', f'PO-{purchase.id:06d}', 'Date:', purchase.created_at.strftime('%d %b %Y')],
            ['Supplier:', purchase.supplier.name, 'Phone:', purchase.supplier.phone or '—'],
            ['Added By:', purchase.added_by.username if purchase.added_by else '—',
             'Status:', 'Paid' if purchase.balance_due == 0 else
                        'Partially Paid' if purchase.amount_paid > 0 else 'Unpaid'],
        ], colWidths=[3.5*cm, 7*cm, 3*cm, 5*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph('Items Purchased', heading))

        items     = purchase.items.select_related('product')
        item_data = [['#', 'Product', 'Barcode', 'Qty', 'Unit Cost (Rs)', 'Total (Rs)']]
        for i, item in enumerate(items, 1):
            item_data.append([str(i), item.product_name,
                               item.product.barcode if item.product else '—',
                               str(item.quantity), f'{item.unit_cost:.2f}',
                               f'{item.total_cost:.2f}'])
        if not items.exists():
            item_data.append(['—', 'No items recorded', '', '', '', ''])

        item_table = Table(item_data, colWidths=[1*cm, 7*cm, 3.5*cm, 1.5*cm, 3.5*cm, 3*cm])
        item_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8f9fa'), colors.white]),
            ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
            ('PADDING',    (0,0), (-1,-1), 7),
            ('ALIGN',      (3,0), (-1,-1), 'RIGHT'),
        ]))
        story.append(item_table)
        story.append(Spacer(1, 0.4*cm))

        totals = Table([
            ['', 'Total Amount:',  f'Rs {purchase.total_amount:.2f}'],
            ['', 'Amount Paid:',   f'Rs {purchase.amount_paid:.2f}'],
            ['', 'Balance Due:',   f'Rs {purchase.balance_due:.2f}'],
        ], colWidths=[9*cm, 5*cm, 5*cm])
        totals.setStyle(TableStyle([
            ('FONTNAME', (1,0), (1,-1), 'Helvetica-Bold'),
            ('ALIGN',    (1,0), (-1,-1), 'RIGHT'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('FONTNAME', (1,2), (-1,2), 'Helvetica-Bold'),
            ('TEXTCOLOR',(2,2),(2,2),
             colors.HexColor('#cc0000') if purchase.balance_due > 0 else colors.HexColor('#007700')),
            ('LINEABOVE',(1,2),(-1,2), 1, colors.black),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(totals)
        if purchase.notes:
            story += [Spacer(1, 0.4*cm), Paragraph(f'<b>Notes:</b> {purchase.notes}', normal)]
        story += [
            Spacer(1, 1*cm),
            HRFlowable(width='100%', thickness=0.5, color=colors.grey),
            Paragraph(f'Generated by SmartPOS | {store_name}',
                      ParagraphStyle('foot', fontSize=8, textColor=colors.grey,
                                     alignment=1, spaceBefore=4)),
        ]
        doc.build(story)
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/pdf')
        response['Content-Disposition'] = \
            f'attachment; filename="purchase_invoice_PO{purchase.id:06d}.pdf"'
        return response

    except ImportError:
        messages.error(request, 'reportlab required. Run: pip install reportlab')
        return redirect('supplier_payments', supplier_id=purchase.supplier.id)


# ══════════════════════════════════════════════════════════════════════════════
# WHATSAPP
# ══════════════════════════════════════════════════════════════════════════════

@store_required
def send_whatsapp(request, sale_id):
    if request.user.is_superuser:
        sale = get_object_or_404(Sale, id=sale_id)
    else:
        sale = get_object_or_404(Sale, id=sale_id, store=request.store)
    phone = request.POST.get('phone', '').strip() if request.method == 'POST' else ''
    if not phone:
        phone = sale.customer.phone if sale.customer else ''
    if not phone:
        return JsonResponse({'success': False, 'message': 'No phone number provided.'})

    # FIX SEC-PHONE: Strict numeric-only phone validation
    phone = _re.sub(r'[\s\-\(\)\+]', '', phone)
    if not _re.match(r'^\d{7,15}$', phone):
        return JsonResponse({'success': False, 'message': 'Invalid phone number.'})

    if phone.startswith('0'):
        phone = '92' + phone[1:]
    elif not phone.startswith('92') and len(phone) == 10:
        phone = '92' + phone

    store_name = request.store.name
    lines = [
        f"🧾 *Receipt from {store_name}*",
        f"Receipt #: *{sale.sale_number}*",
        f"Date: {sale.created_at.strftime('%d %b %Y, %H:%M')}",
    ]
    if sale.customer:
        lines.append(f"Customer: {sale.customer.name}")
    lines += ["", "*Items:*"]
    for item in sale.items.all():
        lines.append(f"  • {item.product_name} x{item.quantity} = Rs {item.total_price:.0f}")
    lines += [
        "",
        f"Subtotal: Rs {sale.subtotal:.0f}",
    ]
    if sale.discount_amount > 0:
        lines.append(f"Discount: -Rs {sale.discount_amount:.0f}")
    if sale.tax_amount > 0:
        lines.append(f"Tax: Rs {sale.tax_amount:.0f}")
    lines.append(f"*TOTAL: Rs {sale.total_amount:.0f}*")
    lines.append(f"Payment: {sale.get_payment_method_display()}")
    if sale.payment_method == 'cash':
        lines.append(f"Cash: Rs {sale.amount_received:.0f} | Change: Rs {sale.change_amount:.0f}")
    if sale.loyalty_points_earned > 0:
        lines.append(f"⭐ Points Earned: {sale.loyalty_points_earned}")
    lines += ["", "Thank you for shopping with us! 🙏"]

    import urllib.parse
    wa_url = f"https://wa.me/{phone}?text={urllib.parse.quote(chr(10).join(lines))}"
    return JsonResponse({'success': True, 'wa_url': wa_url, 'message': 'Ready to send!'})


@store_required
@role_required('admin', 'manager')
def whatsapp_settings(request):
    store        = request.store
    settings_obj, _ = StoreSettings.objects.get_or_create(store=store)

    if request.method == 'POST':
        settings_obj.whatsapp_enabled       = 'whatsapp_enabled' in request.POST
        settings_obj.whatsapp_token         = request.POST.get('whatsapp_token', '').strip()
        settings_obj.whatsapp_phone_id      = request.POST.get('whatsapp_phone_id', '').strip()
        settings_obj.whatsapp_template_name = request.POST.get('whatsapp_template_name', 'receipt').strip()
        settings_obj.save()
        messages.success(request, 'WhatsApp settings saved.')
        return redirect('whatsapp_settings')

    return render(request, 'pos/whatsapp_settings.html', {
        'store': store, 'settings': settings_obj,
    })


# ══════════════════════════════════════════════════════════════════════════════
# CLOUD BACKUP
# ══════════════════════════════════════════════════════════════════════════════

@store_required
@role_required('admin', 'manager')
def cloud_backup_settings(request):
    store        = request.store
    settings_obj, _ = StoreSettings.objects.get_or_create(store=store)

    if request.method == 'POST':
        action = request.POST.get('action', 'save')

        if action == 'test':
            creds = request.POST.get('gdrive_credentials_json', '').strip()
            if not creds:
                return JsonResponse({'ok': False, 'error': 'No credentials pasted.'})
            try:
                svc   = _get_drive_service(creds)
                about = svc.about().get(fields='user').execute()
                return JsonResponse({'ok': True,
                                     'email': about.get('user', {}).get('emailAddress', 'unknown')})
            except Exception as exc:
                logger.error(f"GDrive test error: {exc}")
                return JsonResponse({'ok': False, 'error': 'Connection test failed.'})

        elif action == 'manual_backup':
            if not settings_obj.gdrive_credentials_json:
                messages.error(request, 'No credentials saved. Save your settings first.')
                return redirect('cloud_backup')
            threading.Thread(
                target=_gdrive_upload_excel_async,
                args=(settings_obj.id, 'manual'),
                daemon=True,
            ).start()
            messages.success(request, 'Full backup uploading to Google Drive in background.')
            return redirect('cloud_backup')

        else:
            settings_obj.gdrive_credentials_json = request.POST.get('gdrive_credentials_json', '').strip()
            settings_obj.gdrive_folder_id        = request.POST.get('gdrive_folder_id', '').strip()
            settings_obj.gdrive_enabled          = request.POST.get('gdrive_enabled') == 'on'
            settings_obj.gdrive_backup_on_sale   = request.POST.get('gdrive_backup_on_sale') == 'on'
            settings_obj.gdrive_backup_on_expense= request.POST.get('gdrive_backup_on_expense') == 'on'
            settings_obj.save()
            audit(request, 'settings_change', 'Cloud Backup settings updated')
            messages.success(request, 'Cloud Backup settings saved.')
            return redirect('cloud_backup')

    return render(request, 'pos/cloud_backup.html', {'s': settings_obj})


# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE DRIVE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _get_drive_service(credentials_json_str: str):
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    info  = json.loads(credentials_json_str)
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=['https://www.googleapis.com/auth/drive.file']
    )
    return build('drive', 'v3', credentials=creds, cache_discovery=False)


def _gdrive_upload_excel_async(settings_obj_id: int, label: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from googleapiclient.http import MediaIoBaseUpload
        from pos.models import StoreSettings, Sale, SaleItem, Product, Expense, Customer

        settings_obj = StoreSettings.objects.get(id=settings_obj_id)
        if not settings_obj.gdrive_credentials_json:
            return

        store = settings_obj.store
        wb    = openpyxl.Workbook()
        hdr_font  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        hdr_fill  = PatternFill('solid', start_color='1a1a2e')
        hdr_align = Alignment(horizontal='center', vertical='center')
        fill_a    = PatternFill('solid', start_color='F8F9FA')
        fill_b    = PatternFill('solid', start_color='FFFFFF')

        def style_sheet(ws, headers, rows):
            ws.append(headers)
            for cell in ws[1]:
                cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align
            ws.row_dimensions[1].height = 20
            for i, row in enumerate(rows, start=2):
                ws.append(row)
                fill = fill_a if i % 2 == 0 else fill_b
                for cell in ws[i]:
                    cell.fill = fill; cell.font = Font(name='Arial', size=9)
                    cell.alignment = Alignment(vertical='center')
            for col_idx, _ in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                max_len = max((len(str(ws.cell(r, col_idx).value or ''))
                               for r in range(1, ws.max_row + 1)), default=10)
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        ws1 = wb.active; ws1.title = 'Sales'
        sales_qs = Sale.objects.filter(store=store).select_related(
            'cashier', 'customer').order_by('-created_at')
        style_sheet(ws1,
            ['Sale #','Date','Time','Cashier','Customer','Subtotal','Discount','Tax',
             'Total','Payment','Cash Received','Change','Pts Earned','Pts Used','Status'],
            [[s.sale_number, s.created_at.strftime('%d/%m/%Y'), s.created_at.strftime('%H:%M'),
              s.cashier.username if s.cashier else '-',
              s.customer.name if s.customer else 'Walk-in',
              float(s.subtotal), float(s.discount_amount), float(s.tax_amount),
              float(s.total_amount), s.get_payment_method_display(),
              float(s.amount_received), float(s.change_amount),
              s.loyalty_points_earned, s.loyalty_points_used, s.status.title()]
             for s in sales_qs]
        )
        ws2 = wb.create_sheet('Products')
        style_sheet(ws2,
            ['Name','Barcode','Category','Sell Price','Cost Price','Stock','Supplier'],
            [[p.name, p.barcode, p.category.name if p.category else '-',
              float(p.price), float(p.cost_price), p.stock_quantity,
              p.supplier.name if p.supplier else '-']
             for p in Product.objects.filter(store=store, is_active=True).select_related('category','supplier')]
        )
        ws3 = wb.create_sheet('Expenses')
        style_sheet(ws3,
            ['Date','Category','Title','Amount','Added By'],
            [[str(e.date), str(e.category) if e.category else '-', e.title, float(e.amount),
              e.added_by.username if e.added_by else '-']
             for e in Expense.objects.filter(store=store).order_by('-date')]
        )
        ws4 = wb.create_sheet('Customers')
        style_sheet(ws4,
            ['Name','Phone','Email','Loyalty Points','Joined'],
            [[c.name, c.phone, c.email, c.loyalty_points, c.created_at.strftime('%d/%m/%Y')]
             for c in Customer.objects.filter(store=store)]
        )

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        service  = _get_drive_service(settings_obj.gdrive_credentials_json)
        ts       = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'smartpos_backup_{label}_{ts}.xlsx'
        file_meta = {'name': filename}
        if settings_obj.gdrive_folder_id:
            file_meta['parents'] = [settings_obj.gdrive_folder_id]
        media    = MediaIoBaseUpload(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        uploaded = service.files().create(
            body=file_meta, media_body=media, fields='id,webViewLink'
        ).execute()
        settings_obj.gdrive_last_backup_url = uploaded.get('webViewLink', '')
        settings_obj.gdrive_last_backup_at  = timezone.now()
        settings_obj.save(update_fields=['gdrive_last_backup_url', 'gdrive_last_backup_at'])
    except Exception as exc:
        logger.error(f'[GDrive] async backup error: {exc}')


def _trigger_gdrive_auto_backup(store, trigger_label: str, check_field: str):
    try:
        from pos.models import StoreSettings
        if not store:
            return
        settings_obj, _ = StoreSettings.objects.get_or_create(store=store)
        if not settings_obj.gdrive_enabled:
            return
        if not settings_obj.gdrive_credentials_json:
            return
        if not getattr(settings_obj, check_field, False):
            return
        threading.Thread(
            target=_gdrive_upload_excel_async,
            args=(settings_obj.id, trigger_label),
            daemon=True,
        ).start()
    except Exception as exc:
        logger.error(f'[GDrive] trigger error: {exc}')